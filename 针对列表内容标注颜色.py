import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 读取 A 文件的 A 列内容,要加颜色的内容
df_a = pd.read_excel('数据.xlsx', sheet_name=0)
keywords = df_a.iloc[:, 0].dropna().tolist()

# 加载 B 文件的工作簿和工作表,匹配中A的内容,就在B这里高亮
wb_b = load_workbook('大兴.xlsx')
sheet_b = wb_b.active

# 创建一个 PatternFill 对象，用于设置单元格颜色为浅绿色
fill = PatternFill(fill_type="solid", fgColor="ffff00")

# 遍历 B 文件的每个单元格
for row in sheet_b.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str):#是一个常用的条件检查，用于确保单元格的值存在且为字符串类型
            # 遍历关键词列表，构建正则表达式
            for keyword in keywords:
                # 构建正则表达式，前后加空格
                pattern = re.compile(rf".* {re.escape(keyword)} .*", re.IGNORECASE)
                if pattern.search(cell.value):
                    print(f"匹配成功: {cell.value}，关键字: {keyword}")
                    cell.fill = fill  # 设置单元格颜色为浅绿色
sheet_b['A5'].fill = PatternFill(start_color="FFFF0000", fill_type="solid")
# 保存修改后的内容到新的文件
output_file = 'B_1.xlsx'
wb_b.save(output_file)
print(f"Modified content saved to {output_file}")